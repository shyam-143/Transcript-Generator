import csv
import os
import shutil
import openpyxl
import sys

if(not os.path.exists(r"marksheets")):
    os.mkdir(r"marksheets")
lst_to_remove = os.listdir("marksheets")
for i in lst_to_remove:
    if(i != "concise_marksheet.xlsx"):
        os.remove(f"marksheets//{i}")
if not os.path.exists(r"public//sample_input//master_roll.csv"):
    print("please upload master_roll.csv file")
    exit()
if not os.path.exists(r"public//sample_input//responses.csv"):
    print("please upload responses.csv file")
    exit()
with open(r"public//sample_input//master_roll.csv", 'r') as file:
    rows = csv.reader(file)
    MR = {line[0]: [line[1], 0, 0, 0] for line in rows if line[0] != "roll"}
    m_lst = sorted(MR.keys())
with open(r"public//sample_input//responses.csv", 'r') as file:
    rows = csv.reader(file)
    Options_Name = {option[6].upper(): [element for element in option[7:]]
                    for option in rows if (option[6] != 'Roll Number' and option[6] != "")}
    try:
        i_p, i_n, Final_Score, Final_Count = float(
            sys.argv[1]), float(sys.argv[2]), [], []
    except:
        print("Please enter valid input")
        exit()
    for Roll, options in Options_Name.items():
        Right, Wrong, Not_Attempt = 0, 0, 0
        for i, option in enumerate(options):
            try:
                if (option == Options_Name["ANSWER"][i]):
                    Right += 1
                elif(option):
                    Wrong += 1
                else:
                    Not_Attempt += 1
            except:
                print("Please provide ANSWER in the responses.csv file")
                exit()
        Final_Count.append(f"[{Right},{Wrong},{Not_Attempt}]")
        Final_Score.append(f"{Right*i_p+Wrong*i_n}/140")
        try:
            MR[f"{Roll}"][1], MR[f"{Roll}"][2], MR[f"{Roll}"][3] = Right, Wrong, Not_Attempt
        except:
            pass


img = openpyxl.drawing.image.Image("public//Title.png")
right = openpyxl.styles.Alignment(horizontal='right')
center = openpyxl.styles.Alignment(horizontal='center')
bd = openpyxl.styles.Side(style='thin', color="000000")
element = openpyxl.styles.Font(name='Century', size=12)
thick = openpyxl.styles.Font(name='Century', size=12, bold=True)
red = openpyxl.styles.Font(name='Century', size=12, color="ff0000")
green = openpyxl.styles.Font(name='Century', size=12, color="008000")
blue = openpyxl.styles.Font(name='Century', size=12, color="0000ff")
black = openpyxl.styles.Font(name='Century', size=12, color="000000")
highlight = openpyxl.styles.Border(left=bd, top=bd, right=bd, bottom=bd)
for key in Options_Name.keys():
    wb = openpyxl.Workbook()
    ws = wb["Sheet"]
    ws.title = "quiz"
    ws.add_image(img)
    ws.merge_cells('A5:E5')
    ws["A5"] = "Mark Sheet"
    ws["A5"].alignment = center
    ws["A5"].font = openpyxl.styles.Font(
        size=18, bold=True, name='Century', underline="single")
    for i in [0, 1, 2, 3, 4]:
        ws.column_dimensions[chr(ord("A")+i)].width = 18
    ws.append(["Name:", f"{MR[key][0]}", "", "Exam:", "quiz"])
    ws.append(["Roll Numer:", f"{key}", "", "", ""])
    ws.merge_cells('B6:C6')
    ws.append([])
    ws.append(["", "Right", "Wrong", "Not Attempt", "Max"])
    ws.append(["No.", f"{MR[key][1]}", f"{MR[key][2]}",
              f"{MR[key][3]}", f"{sum(MR[key][1:])}"])
    ws.append(["Marking", f"{i_p}", f"{i_n}", f"{0}", ""])
    ws.append(["Total", f"{MR[key][1]*i_p}", f"{MR[key][2]*i_n}",
              "", f"{MR[key][1]*i_p+MR[key][2]*i_n}/{sum(MR[key][1:])*i_p}"])
    for row in ws.iter_rows(min_row=9, max_row=12):
        for i, cell in enumerate(row):
            cell.font, cell.border, cell.alignment = thick, highlight, center
    ws["B6"].font, ws["B7"].font, ws["E6"].font = thick, thick, thick
    ws["C10"].font, ws["C11"].font, ws["C12"].font = red, red, red
    ws["E12"].font, ws["E13"].font, ws["E14"].font = blue, blue, blue
    ws["B10"].font, ws["B11"].font, ws["B12"].font = green, green, green
    ws["D10"].font, ws["D11"].font, ws["E10"].font = black, black, black
    ws["A6"].font, ws["A7"].font, ws["D6"].font, = element, element, element
    ws["A6"].alignment, ws["A7"].alignment, ws["D6"].alignment, = right, right, right
    ws.append(["Student Ans", "Correct Ans", "", "Student Ans", "Correct Ans"])
    for row in ws.iter_rows(min_row=15, max_row=15):
        for i, cell in enumerate(row):
            if(i != 2):
                cell.font, cell.border, cell.alignment = thick, highlight, center
    for i, row in enumerate(ws.iter_cols(min_col=1, max_col=5, min_row=16, max_row=40)):
        for j, cell in enumerate(row):
            if(i == 1 or i == 4):
                if ((j+24*(i//3)) < len(Options_Name["ANSWER"])):
                    cell.value, cell.font = Options_Name["ANSWER"][j+24*(
                        i//3)], blue
                    cell.border, cell.alignment = highlight, center
            if(i == 0 or i == 3):
                if key in Options_Name:
                    if ((j+24*(i//2)) < len(Options_Name[key])):
                        cell.value, cell.font = Options_Name[key][j+24*(
                            i//2)], green
                        cell.border, cell.alignment = highlight, center
                        if cell.value != Options_Name["ANSWER"][j+24*(i//2)]:
                            cell.font = red
    wb.save(f"marksheets//{key}.xlsx")
for key in Options_Name:
    if (key in m_lst):
        m_lst.remove(key)

for key in m_lst:
    wb = openpyxl.Workbook()
    ws = wb["Sheet"]
    ws.title = "quiz"
    ws.add_image(img)
    ws.merge_cells('A5:E5')
    ws["A5"] = "Mark Sheet"
    ws["A5"].alignment = center
    ws["A5"].font = openpyxl.styles.Font(
        size=18, bold=True, name='Century', underline="single")
    for i in [0, 1, 2, 3, 4]:
        ws.column_dimensions[chr(ord("A")+i)].width = 18
    ws.append(["Name:", f"{MR[key][0]}", "", "Exam:", "quiz"])
    ws.append(["Roll Numer:", f"{key}", "", "Status:", "Absent"])
    ws.merge_cells('B6:C6')
    ws["B6"].font, ws["B7"].font, ws["E6"].font, ws["E7"].font = thick, thick, thick, red
    ws["A6"].font, ws["A7"].font, ws["D6"].font, ws["D7"].font = element, element, element, element
    ws["A6"].alignment, ws["A7"].alignment, ws["D6"].alignment, ws["D7"].alignment = right, right, right, right
    wb.save(f"marksheets//{key}.xlsx")

shutil.make_archive("marksheets", 'zip', "marksheets")
print("Successfully generated roll number wise marksheets")
